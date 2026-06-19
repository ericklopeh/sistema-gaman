from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl

from app.core.config import settings
from app.services.excel_helpers import (
    escribir_seguro,
    estilo_datos,
    limpiar_external_links,
    parse_money,
    safe_name,
)


def arreglar_formula_final_sindicato(ws):
    ws["G13"].value = (
        "=INDEX(Hoja2!A2:HH40000,"
        "MATCH(Hoja1!F13,Hoja2!A2:A40000,0),"
        "MATCH(Hoja1!F14,Hoja2!A2:HH2,0))"
    )


def generar_sindicato_bytes(data: dict, template_path: Path | None = None) -> tuple[bytes, str]:
    """Genera hoja sindicato usando formato_autorizaciones.xlsx (formato legacy)."""
    template_path = template_path or (
        Path(settings.STORAGE_PATH) / "templates" / "formato_autorizaciones.xlsx"
    )
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active

    limpiar_external_links(wb)
    arreglar_formula_final_sindicato(ws)

    estilo, alineacion = estilo_datos()

    def escribir(celda, valor, centrar=True):
        escribir_seguro(
            ws, celda, valor, font=estilo,
            alignment=alineacion if centrar else None,
        )

    def escribir_money(celda, valor_float):
        escribir_seguro(
            ws, celda, float(valor_float), font=estilo,
            alignment=alineacion, number_format='"$"#,##0.00',
        )

    cliente = data["cliente"]
    folio = str(data["folio"])
    fecha_txt = data["fecha"]
    inicio = data["inicio"]
    anio = inicio.split("-")[1] if "-" in inicio else str(datetime.now().year)

    escribir("D1", data.get("telefono", ""))
    escribir("B2", cliente, centrar=False)
    escribir("E2", data.get("rfc", ""))

    f_obj = datetime.strptime(fecha_txt, "%d/%m/%Y")
    escribir_seguro(ws, "G2", f_obj, font=estilo, alignment=alineacion, number_format="dd/mm/yyyy")

    productos = data.get("productos", [])
    for i in range(5):
        r = 4 + i
        prod = productos[i] if i < len(productos) else {}
        escribir(f"A{r}", prod.get("producto", ""), centrar=False)
        escribir(f"B{r}", prod.get("producto_secundario", ""), centrar=False)

        trans = prod.get("trans_credito")
        credito = prod.get("credito")
        precio = prod.get("precio_venta")
        descuento = prod.get("descuento")

        transferencia = trans if trans is not None else precio
        costo = precio if precio is not None else transferencia
        efectivo = None
        if precio is not None and descuento is not None:
            efectivo = max(0.0, float(precio) - float(descuento))
        elif precio is not None:
            efectivo = float(precio)

        if transferencia is not None:
            escribir_money(f"C{r}", transferencia)
        if credito is not None:
            escribir_money(f"D{r}", credito)
        if costo is not None:
            escribir_money(f"E{r}", costo)
        if efectivo is not None:
            escribir_money(f"F{r}", efectivo)

        tipo = prod.get("tipo_vta", "")
        escribir(f"G{r}", tipo)

    escribir("A11", data.get("observaciones", ""), centrar=False)
    escribir("B15", data.get("vendedor", ""), centrar=False)
    escribir("D11", folio)
    escribir("F11", int(data["semana"]))
    escribir("G11", int(anio))
    escribir_seguro(ws, "F13", inicio, font=estilo, alignment=alineacion, number_format="@")
    escribir_seguro(ws, "F14", int(data["plazo"]), font=estilo, alignment=alineacion, number_format="0")

    monto_capt = data.get("monto_total")
    if monto_capt is None:
        total = 0.0
        for prod in productos[:5]:
            pv = parse_money(prod.get("precio_venta"))
            desc = parse_money(prod.get("descuento"))
            if pv in (None, "INVALID"):
                continue
            if desc in (None, "INVALID"):
                desc = 0.0
            total += pv - desc
        monto_capt = total

    escribir_money("D12", monto_capt)

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    filename = f"SINDICATO_{safe_name(folio)}_{safe_name(cliente)}.xlsx"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), filename