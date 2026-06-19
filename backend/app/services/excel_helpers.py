import os
import re
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter


FUENTE_GLOBAL = "Aptos Narrow"
TAM_GLOBAL = 12


def safe_name(texto: str) -> str:
    texto = (texto or "").strip()
    return "".join(c for c in texto if c not in '<>:"/\\|?*').strip()


def limpiar_nombre_archivo(texto: str) -> str:
    texto = texto.upper().strip()
    for caracter in ['\\', '/', ':', '*', '?', '"', '<', '>', '|']:
        texto = texto.replace(caracter, "")
    return texto.replace(" ", "_")


def parse_money(value) -> float | str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    s = s.replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return "INVALID"


def parse_int(value) -> int | str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    s = s.replace(",", "")
    try:
        return int(float(s))
    except ValueError:
        return "INVALID"


def validar_inicio_mm_aaaa(s: str) -> bool:
    return bool(re.match(r"^(0[1-9]|1[0-2])-\d{4}$", (s or "").strip()))


def validar_fecha_ddmmyyyy(s: str) -> bool:
    try:
        datetime.strptime((s or "").strip(), "%d/%m/%Y")
        return True
    except ValueError:
        return False


def guardar_workbook_seguro(wb, ruta_final: Path) -> Path:
    ruta_final = Path(ruta_final)
    ruta_final.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, str(ruta_final))
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except OSError:
            pass
    return ruta_final


def top_left_merge(ws, cell_addr: str) -> str:
    row, col = coordinate_to_tuple(cell_addr)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return cell_addr


def escribir_seguro(ws, celda, valor, font=None, alignment=None, number_format=None):
    celda = top_left_merge(ws, celda)
    cell = ws[celda]
    cell.value = valor
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format


def limpiar_external_links(wb):
    try:
        wb._external_links = []
    except Exception:
        pass


def estilo_datos():
    return (
        Font(name=FUENTE_GLOBAL, size=TAM_GLOBAL, bold=True, color="000000"),
        Alignment(horizontal="center", vertical="center"),
    )