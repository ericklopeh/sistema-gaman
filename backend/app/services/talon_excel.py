from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.config import settings
from app.services.excel_helpers import limpiar_nombre_archivo


def escribir_celda(ws, celda: str, valor):
    for rango in ws.merged_cells.ranges:
        if celda in rango:
            celda_principal = rango.coord.split(":")[0]
            ws[celda_principal] = valor
            return
    ws[celda] = valor


def detalle_cuentas(cuentas_terminadas: list[dict], sumar_a_liquidez: bool) -> str:
    detalles = []
    for cuenta in cuentas_terminadas:
        if bool(cuenta.get("sumar_a_liquidez", False)) != sumar_a_liquidez:
            continue
        qna = str(cuenta.get("qna_termina", "")).strip() or "Sin QNA"
        monto = float(cuenta.get("saldo_liberado", 0) or 0)
        observacion = str(cuenta.get("observacion", "")).strip()
        detalle = f"{qna}: ${monto:,.2f}"
        if observacion:
            detalle += f" | {observacion}"
        detalles.append(detalle)
    return "\n".join(detalles)


def agregar_hoja_cuentas_terminadas(wb, datos: dict, revision: dict, promotor: str, qna: str):
    nombre_hoja = "CUENTAS_TERMINADAS"
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]

    ws = wb.create_sheet(nombre_hoja)
    encabezados = [
        "fecha_revision", "cliente", "rfc", "vendedor", "qna_revision",
        "qna_termina", "saldo_liberado", "sumar_a_liquidez", "observacion",
    ]
    ws.append(encabezados)
    fecha_revision = datetime.now().strftime("%d/%m/%Y")

    for cuenta in revision.get("cuentas_terminadas", []):
        ws.append([
            fecha_revision,
            datos["nombre"],
            datos["rfc"],
            promotor,
            qna,
            cuenta.get("qna_termina", ""),
            float(cuenta.get("saldo_liberado", 0) or 0),
            "Sí" if cuenta.get("sumar_a_liquidez", False) else "No",
            cuenta.get("observacion", ""),
        ])

    color_encabezado = "1F4E78"
    for celda in ws[1]:
        celda.font = Font(color="FFFFFF", bold=True)
        celda.fill = PatternFill("solid", fgColor=color_encabezado)
        celda.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    for celda in ws["G"][1:]:
        celda.number_format = '$#,##0.00;-$#,##0.00;$-'


def generar_excel_revision_bytes(
    datos: dict,
    revision: dict,
    mensaje_vendedor: str,
    promotor: str,
    qna: str = "09-2026",
    template_path: Path | None = None,
) -> tuple[bytes, str]:
    template_path = template_path or Path(settings.STORAGE_PATH) / "templates" / "plantilla_revision_talon.xlsx"
    wb = load_workbook(template_path)
    ws = wb["Hoja1"]
    codigos = revision["codigos_revision"]

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    escribir_celda(ws, "B3", datos["rfc"])
    escribir_celda(ws, "D3", datos["nombre"])
    escribir_celda(ws, "C4", qna)
    escribir_celda(ws, "E4", datos.get("fecha_pago", ""))

    for codigo, fila in [
        ("E4", "B6"), ("E3", "B7"), ("Q", "B8"), ("CP", "B9"), ("7", "B10"),
        ("CT", "B11"), ("7B", "B12"), ("E9", "B13"), ("SG", "B14"), ("O1", "B15"),
    ]:
        escribir_celda(ws, fila, codigos[codigo])

    escribir_celda(ws, "C6", revision["ingresos"])
    escribir_celda(ws, "E6", revision["descuentos"])
    escribir_celda(ws, "E7", revision["saldo_100"])
    escribir_celda(ws, "E8", revision["total_para_venta_70"])
    escribir_celda(ws, "E10", revision["saldo_70"])
    escribir_celda(ws, "D12", revision["saldo_70"])
    escribir_celda(ws, "D15", revision["saldo_100"])
    escribir_celda(ws, "B16", codigos["DC"])
    escribir_celda(ws, "D16", codigos["DC"])
    escribir_celda(ws, "A18", "")

    for celda in ["C20", "C21", "C22", "C23", "E20", "E21", "E22", "E23", "B25", "D25"]:
        escribir_celda(ws, celda, 0)

    escribir_celda(ws, "B28", revision["saldo_70"])
    escribir_celda(ws, "D28", revision["saldo_100"])
    escribir_celda(ws, "B31", promotor)
    escribir_celda(ws, "B32", datos["nombre"])
    escribir_celda(ws, "B33", datos["rfc"])
    escribir_celda(ws, "B34", revision["liquidez_final"])
    escribir_celda(ws, "B35", qna)
    escribir_celda(ws, "B36", mensaje_vendedor)
    escribir_celda(ws, "B40", datetime.now().strftime("%d/%m/%Y"))

    cuentas_terminadas = revision.get("cuentas_terminadas", [])
    detalle_liberadas = detalle_cuentas(cuentas_terminadas, True)
    detalle_observadas = detalle_cuentas(cuentas_terminadas, False)

    for merge_range in [
        "A41:C41", "D41:E41", "A42:C42", "D42:E42", "A43:C43", "D43:E43",
        "A44:C44", "D44:E44", "A45:C45", "D45:E45", "A46:C46", "D46:E46",
        "A47:C47", "D47:E47", "A48:E48", "A49:E51", "A52:E52", "A53:E55",
    ]:
        ws.merge_cells(merge_range)

    escribir_celda(ws, "A41", "LIQUIDEZ DEL TALÓN")
    escribir_celda(ws, "D41", revision.get("liquidez_talon", revision["saldo_100"]))
    escribir_celda(ws, "A42", "APOYO ADICIONAL")
    escribir_celda(ws, "D42", revision.get("abono_extra", 0))
    escribir_celda(ws, "A43", "TOTAL SALDO LIBERADO")
    escribir_celda(ws, "D43", revision.get("total_saldo_liberado", 0))
    escribir_celda(ws, "A44", "TOTAL SOLO OBSERVADO")
    escribir_celda(ws, "D44", revision.get("total_solo_observado", 0))
    escribir_celda(ws, "A45", "MONTO PROGRAMADO")
    escribir_celda(ws, "D45", revision.get("programado", 0))
    escribir_celda(ws, "A46", "LIQUIDEZ ANTES LIBERACIÓN")
    escribir_celda(ws, "D46", revision.get("liquidez_antes_liberacion", revision["liquidez_final"]))
    escribir_celda(ws, "A47", "LIQUIDEZ FINAL")
    escribir_celda(ws, "D47", revision["liquidez_final"])
    escribir_celda(ws, "A48", "DETALLE CUENTAS LIBERADAS")
    escribir_celda(ws, "A49", detalle_liberadas or "Sin cuentas sumadas a liquidez")
    escribir_celda(ws, "A52", "DETALLE CUENTAS OBSERVADAS")
    escribir_celda(ws, "A53", detalle_observadas or "Sin cuentas solo observadas")

    for celda in ["A41", "A42", "A43", "A44", "A45", "A46", "A47", "A48", "A52"]:
        ws[celda].font = Font(bold=True)

    ws["A49"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A53"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.print_area = "A1:E55"

    celdas_moneda = [
        "B6", "B7", "B8", "B9", "B10", "B11", "B12", "B13", "B14", "B15",
        "C6", "E6", "E7", "E8", "E10", "D12", "D15", "B16", "D16",
        "B25", "D25", "B28", "D28", "B34", "D41", "D42", "D43", "D44", "D45", "D46", "D47",
    ]
    for celda in celdas_moneda:
        ws[celda].number_format = '$#,##0.00;-$#,##0.00;$-'

    agregar_hoja_cuentas_terminadas(wb, datos, revision, promotor, qna)

    nombre_cliente = limpiar_nombre_archivo(datos["nombre"])
    nombre_promotor = limpiar_nombre_archivo(promotor)
    filename = f"REVISION_{nombre_cliente}_{nombre_promotor}.xlsx"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), filename