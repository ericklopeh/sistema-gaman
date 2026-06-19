from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl

from app.core.config import settings
from app.services.excel_helpers import (
    FUENTE_GLOBAL,
    TAM_GLOBAL,
    escribir_seguro,
    estilo_datos,
    limpiar_external_links,
    limpiar_nombre_archivo,
    parse_money,
    safe_name,
)


def arreglar_formula_final(ws):
    ws["G13"].value = (
        "=INDEX(Hoja2!A2:CV40000,"
        "MATCH(Hoja1!F13,Hoja2!A2:A40000,0),"
        "MATCH(Hoja1!F14,Hoja2!A1:CV1,0))"
    )


def generar_autorizacion_bytes(data: dict, template_path: Path | None = None) -> tuple[bytes, str]:
    template_path = template_path or (
        Path(settings.STORAGE_PATH) / "templates" / "plantilla_master_autorizaciones.xlsx"
    )
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active

    limpiar_external_links(wb)
    arreglar_formula_final(ws)

    estilo, alineacion = estilo_datos()

    def escribir(celda, valor, centrar=True):
        escribir_seguro(
            ws, celda, valor, font=estilo,
            alignment=alineacion if centrar else None,
        )

    def escribir_money(celda, valor_float):
        escribir_seguro(
            ws, celda, float(valor_float), font=estilo,
            alignment=alineacion, number_format='"$"#,##0.00',
        )

    cliente = data["cliente"]
    folio = str(data["folio"])
    fecha_txt = data["fecha"]

    escribir("D1", data.get("telefono", ""))
    escribir("B2", cliente, centrar=False)
    escribir("E2", data.get("rfc", ""))

    f_obj = datetime.strptime(fecha_txt, "%d/%m/%Y")
    escribir_seguro(ws, "G2", f_obj, font=estilo, alignment=alineacion, number_format="dd/mm/yyyy")

    productos = data.get("productos", [])
    for i in range(5):
        r = 4 + i
        prod = productos[i] if i < len(productos) else {}
        escribir(f"A{r}", prod.get("producto", ""), centrar=False)
        for campo, col in [
            ("trans_credito", "C"), ("credito", "D"),
            ("precio_venta", "E"), ("descuento", "F"),
        ]:
            valor = prod.get(campo)
            if valor is not None:
                escribir_money(f"{col}{r}", valor)
        tipo = prod.get("tipo_vta", "")
        escribir(f"G{r}", tipo)

    escribir("A11", data.get("observaciones", ""), centrar=False)
    escribir("B15", data.get("vendedor", ""), centrar=False)
    escribir("D11", folio)
    escribir("F11", int(data["semana"]))
    escribir_seguro(ws, "F13", data["inicio"], font=estilo, alignment=alineacion, number_format="@")
    escribir_seguro(ws, "F14", int(data["plazo"]), font=estilo, alignment=alineacion, number_format="0")

    monto_capt = data.get("monto_total")
    if monto_capt is None:
        total = 0.0
        for prod in productos[:5]:
            pv = parse_money(prod.get("precio_venta"))
            desc = parse_money(prod.get("descuento"))
            if pv in (None, "INVALID"):
                continue
            if desc in (None, "INVALID"):
                desc = 0.0
            total += pv - desc
        monto_capt = total

    escribir_money("D12", monto_capt)

    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    filename = f"{safe_name(folio)} {safe_name(cliente)}.xlsx"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), filename